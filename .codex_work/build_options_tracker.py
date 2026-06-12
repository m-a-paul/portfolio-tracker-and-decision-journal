from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


SOURCE = Path("/Users/paulmiyake/Downloads/Stocks and Options.xlsx")
OUT_DIR = Path("/Users/paulmiyake/Documents/New project 2/outputs/options_tracker")
OUT_FILE = OUT_DIR / "Options_Trading_Tracker.xlsx"
MAX_TRADE_ROWS = 1000
MAX_POSITION_ROWS = 150
TRACKING_YEAR = 2026

CURRENT_STOCK_POSITIONS = [
    {"ticker": "CASH", "shares": 5526.75, "price": 1.00, "avg_cost": 1.00, "notes": "Estimated free cash / buying power after reconciling known activity through 2026-06-09, the 2026-06-11 SHLS call purchase, and closing SONY calls. Includes the user-reported $3,000 cash addition, which was not visible as a posted deposit in the CSV exports; confirm against broker buying power."},
    {"ticker": "ETH", "shares": 1, "price": 2298.88, "avg_cost": None, "notes": "1 Ethereum token. Cost basis pending."},
    {"ticker": "ADEA", "shares": 100, "price": 29.69, "avg_cost": 29.69, "notes": "Bought 100 shares on 2026-06-08 per joint CSV. Current price set to fill pending live quote refresh."},
    {"ticker": "AEHR", "shares": 200, "price": 99.84, "avg_cost": 35.73, "notes": "Combined holdings from tracker plus YTD CSV activity. Price last refreshed 2026-05-28."},
    {"ticker": "AMD", "shares": 100, "price": 518.15, "avg_cost": 165.46, "notes": "Existing pre-CSV position retained. Price last refreshed 2026-05-28."},
    {"ticker": "BMNR", "shares": 700, "price": 19.25, "avg_cost": 31.58, "notes": "Combined across portfolios; reconciled against YTD activity. Price last refreshed 2026-05-28."},
    {"ticker": "CRWV", "shares": 100, "price": 102.43, "avg_cost": 102.43, "notes": "Bought 50 shares at $103.07 and 50 shares at $101.79 on 2026-06-05; weighted average $102.43."},
    {"ticker": "FIVN", "shares": 100, "price": 23.00, "avg_cost": 23.00, "notes": "Bought 100 shares on 2026-06-09 per joint CSV. Current price set to fill pending live quote refresh."},
    {"ticker": "HOOD", "shares": 775, "price": 84.85, "avg_cost": 41.4806451613, "notes": "Combined across portfolios; internal transfer rows cancel in combined view. Price last refreshed 2026-05-28."},
    {"ticker": "NEXT", "shares": 500, "price": 8.05, "avg_cost": 7.72, "notes": "Combined holdings; internal transfer rows cancel in combined view. Price last refreshed 2026-05-28."},
    {"ticker": "NTNX", "shares": 100, "price": 48.80, "avg_cost": 45.35, "notes": "Bought 100 shares at $45.35 on 2026-05-21. Price last refreshed 2026-05-28."},
    {"ticker": "PLTR", "shares": 770, "price": 143.35, "avg_cost": 18.51, "notes": "Combined across portfolios; reduced by 10 shares sold 2026-06-03 at $144.40 per joint CSV. Price last refreshed 2026-05-28."},
    {"ticker": "TSLA", "shares": 29, "price": 442.15, "avg_cost": 172.94, "notes": "Existing pre-CSV position retained. Price last refreshed 2026-05-28."},
]

CURRENT_OPTION_POSITIONS = [
    {"trade_id": 1, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 4, 24), "close_date": datetime(2026, 5, 26), "ticker": "BMNR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 26.50, "expiration": datetime(2026, 6, 5), "qty": 1, "entry_price": 1.06, "exit_price": 0.10, "fees": 0.04, "notes": "Bought to close per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 2, "strategy": "Cash Secured Put", "status": "Open", "open_date": datetime(2026, 5, 14), "close_date": None, "ticker": "LUMN", "asset": "Option", "action": "Sell", "option_type": "Put", "strike": 11.00, "expiration": datetime(2026, 6, 12), "qty": 1, "entry_price": 1.02, "exit_price": 1.35, "fees": 0, "notes": "Open position retained from tracker; not present in YTD CSV exports.", "source": "Existing tracker"},
    {"trade_id": 3, "strategy": "Cash Secured Put", "status": "Closed", "open_date": datetime(2026, 4, 24), "close_date": datetime(2026, 5, 28), "ticker": "SOFI", "asset": "Option", "action": "Sell", "option_type": "Put", "strike": 17.00, "expiration": datetime(2026, 6, 5), "qty": 1, "entry_price": 0.91, "exit_price": 0.54, "fees": 0.04, "notes": "Bought to close per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 4, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 5, 5), "close_date": datetime(2026, 5, 18), "ticker": "PLTR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 150.00, "expiration": datetime(2026, 6, 5), "qty": 1, "entry_price": 3.55, "exit_price": 1.16, "fees": 0.04, "notes": "Bought to close per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 5, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 5, 6), "close_date": datetime(2026, 5, 28), "ticker": "PLTR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 160.00, "expiration": datetime(2026, 6, 5), "qty": 1, "entry_price": 1.45, "exit_price": 0.21, "fees": 0.04, "notes": "Bought to close per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 6, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 5, 11), "close_date": datetime(2026, 6, 5), "ticker": "PLTR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 175.00, "expiration": datetime(2026, 6, 5), "qty": 1, "entry_price": 0.49, "exit_price": 0.00, "fees": 0, "notes": "Expired per joint CSV option expiration row.", "source": "Broker CSV"},
    {"trade_id": 7, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 5, 4), "close_date": datetime(2026, 5, 26), "ticker": "PLTR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 220.00, "expiration": datetime(2026, 6, 5), "qty": 1, "entry_price": 0.44, "exit_price": 0.03, "fees": 0.04, "notes": "Bought to close per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 8, "strategy": "Cash Secured Put", "status": "Closed", "open_date": datetime(2026, 5, 6), "close_date": datetime(2026, 5, 19), "ticker": "ASTS", "asset": "Option", "action": "Sell", "option_type": "Put", "strike": 65.00, "expiration": datetime(2026, 6, 5), "qty": 1, "entry_price": 5.75, "exit_price": 1.50, "fees": 0.04, "notes": "Bought to close per user/broker activity.", "source": "Broker CSV"},
    {"trade_id": 9, "strategy": "Cash Secured Put", "status": "Closed", "open_date": datetime(2026, 5, 4), "close_date": datetime(2026, 5, 15), "ticker": "NEXT", "asset": "Option", "action": "Sell", "option_type": "Put", "strike": 8.00, "expiration": datetime(2026, 5, 15), "qty": 2, "entry_price": 0.33, "exit_price": 0.00, "fees": 0, "notes": "Expired/closed after 2026-05-15 expiration; retained from tracker with CSV confirming no later close.", "source": "Existing tracker + Broker CSV"},
    {"trade_id": 10, "strategy": "Debit Spread", "status": "Open", "open_date": datetime(2026, 1, 29), "close_date": None, "ticker": "GRAB", "asset": "Option Spread", "action": "Buy", "option_type": "Call", "strike": 10.00, "expiration": datetime(2028, 1, 21), "qty": 180, "entry_price": 0.10, "exit_price": 0.08, "fees": 0, "notes": "Open GRAB $10/$12 call debit spread; YTD CSV confirms 180 paired contracts.", "source": "Broker CSV + existing tracker"},
    {"trade_id": 11, "strategy": "Covered Call", "status": "Open", "open_date": datetime(2026, 4, 17), "close_date": None, "ticker": "PLTR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 190.00, "expiration": datetime(2027, 1, 15), "qty": 1, "entry_price": 14.73, "exit_price": 9.38, "fees": 0, "notes": "Open PLTR $190 covered call per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 12, "strategy": "Debit Spread", "status": "Open", "open_date": datetime(2025, 1, 24), "close_date": None, "ticker": "GRAB", "asset": "Option Spread", "action": "Buy", "option_type": "Call", "strike": 7.50, "expiration": datetime(2027, 1, 15), "qty": 400, "entry_price": 0.29, "exit_price": 0.03, "fees": 0, "notes": "Consolidated GRAB $7.50/$10 call debit spreads. Existing tracker keeps pre-2026 contracts; CSV confirms 60 YTD additions.", "source": "Existing tracker + Broker CSV"},
    {"trade_id": 13, "strategy": "Long Call/Put", "status": "Open", "open_date": datetime(2025, 11, 28), "close_date": None, "ticker": "BMNR", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 75.00, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 4.75, "exit_price": 0.02, "fees": 0, "notes": "Long BMNR $75 call retained from tracker; opened before CSV period.", "source": "Existing tracker"},
    {"trade_id": 14, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 4, 28), "close_date": datetime(2026, 5, 28), "ticker": "AEHR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 150.00, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 2.90, "exit_price": 2.78, "fees": 0.04, "notes": "Bought to close per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 15, "strategy": "Long Call/Put", "status": "Closed", "open_date": datetime(2026, 5, 15), "close_date": datetime(2026, 6, 9), "ticker": "SHLS", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 10.00, "expiration": datetime(2026, 10, 16), "qty": 4, "entry_price": 2.40, "exit_price": 2.3745, "fees": 0, "notes": "Corrected to 4 SHLS calls: buys total about $960.16; sold 2 at $2.45 and 2 at $2.30 on 2026-06-09.", "source": "Broker CSV + user correction"},
    {"trade_id": 19, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 5, 19), "close_date": datetime(2026, 5, 29), "ticker": "PLTR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 150.00, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 2.24, "exit_price": 8.60, "fees": 0.04, "notes": "Bought to close per joint CSV as part of PLTR roll sequence.", "source": "Broker CSV"},
    {"trade_id": 20, "strategy": "Cash Secured Put", "status": "Closed", "open_date": datetime(2026, 5, 19), "close_date": datetime(2026, 5, 29), "ticker": "ASTS", "asset": "Option", "action": "Sell", "option_type": "Put", "strike": 68.00, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 4.24, "exit_price": 0.75, "fees": 0.04, "notes": "Bought to close per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 22, "strategy": "Long Call/Put", "status": "Open", "open_date": datetime(2026, 5, 18), "close_date": None, "ticker": "NTNX", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 47.50, "expiration": datetime(2026, 10, 16), "qty": 2, "entry_price": 7.15, "exit_price": 7.15, "fees": 0, "notes": "Bought 1 at $7.30 and 1 at $7.00 per joint CSV; average $7.15.", "source": "Broker CSV"},
    {"trade_id": 23, "strategy": "Long Call/Put", "status": "Closed", "open_date": datetime(2026, 5, 19), "close_date": datetime(2026, 5, 26), "ticker": "BLDP", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 4.00, "expiration": datetime(2026, 8, 21), "qty": 1, "entry_price": 1.15, "exit_price": 2.35, "fees": 0.08, "notes": "Sold BLDP $4 call for $234.94 proceeds per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 24, "strategy": "Covered Call", "status": "Open", "open_date": datetime(2026, 5, 28), "close_date": None, "ticker": "PLTR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 160.00, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 0.98, "exit_price": 0.98, "fees": 0, "notes": "Open PLTR $160 covered call per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 25, "strategy": "Cash Secured Put", "status": "Open", "open_date": datetime(2026, 5, 28), "close_date": None, "ticker": "SOFI", "asset": "Option", "action": "Sell", "option_type": "Put", "strike": 18.00, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 1.49, "exit_price": 1.49, "fees": 0, "notes": "Open SOFI $18 put per joint CSV; corrected credit to $1.49.", "source": "Broker CSV"},
    {"trade_id": 26, "strategy": "Covered Call", "status": "Open", "open_date": datetime(2026, 5, 28), "close_date": None, "ticker": "AEHR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 160.00, "expiration": datetime(2026, 7, 2), "qty": 1, "entry_price": 4.05, "exit_price": 4.05, "fees": 0, "notes": "Open AEHR $160 7/2 covered call per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 27, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 5, 27), "close_date": datetime(2026, 6, 9), "ticker": "NTNX", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 55.00, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 0.95, "exit_price": 0.70, "fees": 0.08, "notes": "Bought to close NTNX $55 covered call per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 28, "strategy": "Long Call/Put", "status": "Closed", "open_date": datetime(2026, 5, 26), "close_date": datetime(2026, 6, 5), "ticker": "SOFI", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 22.00, "expiration": datetime(2027, 12, 17), "qty": 2, "entry_price": 4.00, "exit_price": 4.05, "fees": 0.19, "notes": "Sold to close 2 SOFI $22 calls at $4.05 per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 29, "strategy": "Long Call/Put", "status": "Open", "open_date": datetime(2026, 5, 28), "close_date": None, "ticker": "FIVN", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 25.00, "expiration": datetime(2026, 10, 16), "qty": 2, "entry_price": 4.30, "exit_price": 4.30, "fees": 0, "notes": "Open FIVN $25 calls per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 30, "strategy": "Long Call/Put", "status": "Closed", "open_date": datetime(2026, 5, 22), "close_date": datetime(2026, 6, 4), "ticker": "SONY", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 30.00, "expiration": datetime(2026, 9, 18), "qty": 6, "entry_price": 0.36, "exit_price": 0.30, "fees": 0.27, "notes": "Sold to close 6 SONY $30 calls at $0.30 per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 31, "strategy": "Long Call/Put", "status": "Closed", "open_date": datetime(2026, 5, 26), "close_date": datetime(2026, 6, 11), "ticker": "SONY", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 22.50, "expiration": datetime(2026, 12, 18), "qty": 5, "entry_price": 2.80, "exit_price": 1.85, "fees": 0, "notes": "Closed all 5 SONY $22.50 12/18/26 calls at $1.85 each on 2026-06-11. User noted lesson: copy trading someone else did not work out.", "source": "User trade update"},
    {"trade_id": 32, "strategy": "Long Call/Put", "status": "Closed", "open_date": datetime(2026, 5, 26), "close_date": datetime(2026, 6, 11), "ticker": "SONY", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 22.50, "expiration": datetime(2027, 1, 15), "qty": 7, "entry_price": 2.77, "exit_price": 2.00, "fees": 0, "notes": "Closed all 7 SONY $22.50 1/15/27 calls at $2.00 each on 2026-06-11. User noted lesson: copy trading someone else did not work out.", "source": "User trade update"},
    {"trade_id": 35, "strategy": "Covered Call", "status": "Open", "open_date": datetime(2026, 5, 29), "close_date": None, "ticker": "PLTR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 220.00, "expiration": datetime(2027, 1, 15), "qty": 1, "entry_price": 12.50, "exit_price": 12.50, "fees": 0, "notes": "Final open PLTR $220 covered call from 2026-05-29 roll sequence per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 36, "strategy": "Cash Secured Put", "status": "Open", "open_date": datetime(2026, 6, 1), "close_date": None, "ticker": "CRNC", "asset": "Option", "action": "Sell", "option_type": "Put", "strike": 12.00, "expiration": datetime(2026, 6, 18), "qty": 2, "entry_price": 1.05, "exit_price": 1.05, "fees": 0, "notes": "Open 2 CRNC $12 cash secured puts per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 37, "strategy": "Long Call/Put", "status": "Open", "open_date": datetime(2026, 6, 1), "close_date": None, "ticker": "CRNC", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 10.00, "expiration": datetime(2026, 8, 21), "qty": 6, "entry_price": 3.83, "exit_price": 3.83, "fees": 0, "notes": "Open CRNC $10 calls: 2 at $4.10, 2 at $3.60, 2 at $3.80; weighted average about $3.83.", "source": "Broker CSV"},
    {"trade_id": 38, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 5, 29), "close_date": datetime(2026, 5, 29), "ticker": "PLTR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 152.50, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 10.02, "exit_price": 10.75, "fees": 0.08, "notes": "Intraday PLTR roll leg from joint CSV.", "source": "Broker CSV"},
    {"trade_id": 39, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 5, 29), "close_date": datetime(2026, 5, 29), "ticker": "PLTR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 220.00, "expiration": datetime(2027, 1, 15), "qty": 1, "entry_price": 10.60, "exit_price": 12.02, "fees": 0.08, "notes": "Intermediate PLTR $220 roll leg from joint CSV.", "source": "Broker CSV"},
    {"trade_id": 40, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 5, 27), "close_date": datetime(2026, 5, 27), "ticker": "HOOD", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 90.00, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 0.98, "exit_price": 0.95, "fees": 0.08, "notes": "Opened and closed HOOD $90 call per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 41, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 5, 29), "close_date": datetime(2026, 6, 1), "ticker": "HOOD", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 120.00, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 0.73, "exit_price": 0.63, "fees": 0.08, "notes": "Opened and closed HOOD $120 call per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 42, "strategy": "Covered Call", "status": "Closed", "open_date": datetime(2026, 4, 16), "close_date": datetime(2026, 6, 9), "ticker": "AEHR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 110.00, "expiration": datetime(2026, 9, 18), "qty": 1, "entry_price": 18.52, "exit_price": 20.25, "fees": 0.08, "notes": "Rolled to AEHR $160 1/15/27; closed per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 43, "strategy": "Covered Call", "status": "Open", "open_date": datetime(2026, 6, 9), "close_date": None, "ticker": "AEHR", "asset": "Option", "action": "Sell", "option_type": "Call", "strike": 160.00, "expiration": datetime(2027, 1, 15), "qty": 1, "entry_price": 22.05, "exit_price": 22.05, "fees": 0, "notes": "Open AEHR $160 1/15/27 covered call from roll per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 44, "strategy": "Cash Secured Put", "status": "Open", "open_date": datetime(2026, 5, 22), "close_date": None, "ticker": "NEXT", "asset": "Option", "action": "Sell", "option_type": "Put", "strike": 8.00, "expiration": datetime(2026, 6, 18), "qty": 1, "entry_price": 0.45, "exit_price": 0.45, "fees": 0, "notes": "Open NEXT $8 put per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 45, "strategy": "Cash Secured Put", "status": "Open", "open_date": datetime(2026, 6, 9), "close_date": None, "ticker": "ADEA", "asset": "Option", "action": "Sell", "option_type": "Put", "strike": 30.00, "expiration": datetime(2026, 7, 17), "qty": 1, "entry_price": 3.60, "exit_price": 3.60, "fees": 0, "notes": "Open ADEA $30 cash secured put per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 46, "strategy": "Long Call/Put", "status": "Open", "open_date": datetime(2026, 6, 5), "close_date": None, "ticker": "ADEA", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 30.00, "expiration": datetime(2026, 9, 18), "qty": 2, "entry_price": 6.60, "exit_price": 6.60, "fees": 0, "notes": "Open ADEA $30 calls per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 47, "strategy": "Long Call/Put", "status": "Open", "open_date": datetime(2026, 6, 5), "close_date": None, "ticker": "GRAB", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 5.00, "expiration": datetime(2027, 12, 17), "qty": 40, "entry_price": 0.59, "exit_price": 0.59, "fees": 0, "notes": "Open GRAB $5 calls per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 48, "strategy": "Long Call/Put", "status": "Open", "open_date": datetime(2026, 6, 3), "close_date": None, "ticker": "CRNC", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 15.00, "expiration": datetime(2027, 1, 15), "qty": 2, "entry_price": 3.30, "exit_price": 3.30, "fees": 0, "notes": "Open CRNC $15 calls per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 49, "strategy": "Long Call/Put", "status": "Open", "open_date": datetime(2026, 6, 3), "close_date": None, "ticker": "FIVN", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 25.00, "expiration": datetime(2026, 8, 21), "qty": 2, "entry_price": 4.00, "exit_price": 4.00, "fees": 0, "notes": "Open FIVN $25 8/21 calls per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 50, "strategy": "Long Call/Put", "status": "Open", "open_date": datetime(2026, 6, 9), "close_date": None, "ticker": "FIVN", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 22.50, "expiration": datetime(2027, 1, 15), "qty": 1, "entry_price": 5.50, "exit_price": 5.50, "fees": 0, "notes": "Open FIVN $22.50 call per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 55, "strategy": "Long Call/Put", "status": "Open", "open_date": datetime(2026, 6, 11), "close_date": None, "ticker": "SHLS", "asset": "Option", "action": "Buy", "option_type": "Call", "strike": 10.00, "expiration": datetime(2027, 1, 15), "qty": 2, "entry_price": 2.80, "exit_price": 2.80, "fees": 0, "notes": "Bought 2 SHLS $10 calls expiring 2027-01-15 at $2.80 each on 2026-06-11.", "source": "User trade update"},
]

CURRENT_STOCK_TRADES = [
    {"trade_id": 16, "strategy": "Stock Sale", "status": "Closed", "open_date": datetime(2026, 5, 15), "close_date": datetime(2026, 5, 15), "ticker": "BMNR", "asset": "Stock", "action": "Sell", "option_type": "", "strike": None, "expiration": None, "qty": 50, "entry_price": 31.58, "exit_price": 20.25, "fees": 0, "notes": "Sold 50 BMNR shares at $20.25 on 2026-05-15. Entry price uses current weighted average basis.", "source": "User trade update"},
    {"trade_id": 17, "strategy": "Stock Buy", "status": "Closed", "open_date": datetime(2026, 5, 15), "close_date": datetime(2026, 5, 28), "ticker": "FIVN", "asset": "Stock", "action": "Buy", "option_type": "", "strike": None, "expiration": None, "qty": 100, "entry_price": 21.83, "exit_price": 22.77, "fees": 0, "notes": "Bought 100 FIVN shares 2026-05-15 and sold 2026-05-28; consolidated from prior tracker and joint CSV.", "source": "Existing tracker + Broker CSV"},
    {"trade_id": 18, "strategy": "Stock Sale", "status": "Closed", "open_date": datetime(2026, 5, 15), "close_date": datetime(2026, 5, 15), "ticker": "PLTR", "asset": "Stock", "action": "Sell", "option_type": "", "strike": None, "expiration": None, "qty": 10, "entry_price": 18.51, "exit_price": 132.2575, "fees": 0, "notes": "Sold 10 PLTR shares on 2026-05-15. Sale price retained from prior tracker.", "source": "Existing tracker"},
    {"trade_id": 21, "strategy": "Stock Sale", "status": "Closed", "open_date": datetime(2026, 5, 19), "close_date": datetime(2026, 5, 19), "ticker": "PLTR", "asset": "Stock", "action": "Sell", "option_type": "", "strike": None, "expiration": None, "qty": 10, "entry_price": 18.51, "exit_price": 133.98, "fees": 0, "notes": "Sold 10 PLTR shares at $133.98 on 2026-05-19.", "source": "User trade update"},
    {"trade_id": 33, "strategy": "Stock Sale", "status": "Closed", "open_date": datetime(2026, 6, 5), "close_date": datetime(2026, 6, 5), "ticker": "GRAB", "asset": "Stock", "action": "Sell", "option_type": "", "strike": None, "expiration": None, "qty": 4000, "entry_price": 5.00, "exit_price": 3.39, "fees": 1.06, "notes": "Sold all 4,000 GRAB shares across three fills at $3.39 on 2026-06-05 per individual CSV.", "source": "Broker CSV"},
    {"trade_id": 34, "strategy": "Stock Buy", "status": "Open", "open_date": datetime(2026, 5, 21), "close_date": None, "ticker": "NTNX", "asset": "Stock", "action": "Buy", "option_type": "", "strike": None, "expiration": None, "qty": 100, "entry_price": 45.35, "exit_price": None, "fees": 0, "notes": "Bought 100 NTNX shares at $45.35 on 2026-05-21 per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 51, "strategy": "Stock Sale", "status": "Closed", "open_date": datetime(2026, 6, 3), "close_date": datetime(2026, 6, 3), "ticker": "PLTR", "asset": "Stock", "action": "Sell", "option_type": "", "strike": None, "expiration": None, "qty": 10, "entry_price": 18.51, "exit_price": 144.40, "fees": 0.0, "notes": "Sold 10 PLTR shares at $144.40 on 2026-06-03 per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 52, "strategy": "Stock Buy", "status": "Open", "open_date": datetime(2026, 6, 8), "close_date": None, "ticker": "ADEA", "asset": "Stock", "action": "Buy", "option_type": "", "strike": None, "expiration": None, "qty": 100, "entry_price": 29.69, "exit_price": None, "fees": 0, "notes": "Bought 100 ADEA shares at $29.69 on 2026-06-08 per joint CSV.", "source": "Broker CSV"},
    {"trade_id": 53, "strategy": "Stock Buy", "status": "Open", "open_date": datetime(2026, 6, 5), "close_date": None, "ticker": "CRWV", "asset": "Stock", "action": "Buy", "option_type": "", "strike": None, "expiration": None, "qty": 100, "entry_price": 102.43, "exit_price": None, "fees": 0, "notes": "Bought 50 CRWV at $103.07 and 50 at $101.79 on 2026-06-05 per individual CSV; weighted average $102.43.", "source": "Broker CSV"},
    {"trade_id": 54, "strategy": "Stock Buy", "status": "Open", "open_date": datetime(2026, 6, 9), "close_date": None, "ticker": "FIVN", "asset": "Stock", "action": "Buy", "option_type": "", "strike": None, "expiration": None, "qty": 100, "entry_price": 23.00, "exit_price": None, "fees": 0, "notes": "Bought 100 FIVN shares at $23.00 on 2026-06-09 per joint CSV.", "source": "Broker CSV"},
]


COLORS = {
    "navy": "1F4E78",
    "blue": "5B9BD5",
    "pale_blue": "DDEBF7",
    "green": "70AD47",
    "pale_green": "E2F0D9",
    "orange": "F4B183",
    "pale_orange": "FCE4D6",
    "gray": "F2F2F2",
    "dark_gray": "595959",
    "red": "C00000",
    "pale_red": "F4CCCC",
    "white": "FFFFFF",
    "black": "000000",
}


thin_gray = Side(style="thin", color="D9E2F3")
medium_blue = Side(style="medium", color=COLORS["navy"])


def month_start(value):
    if isinstance(value, datetime):
        return datetime(value.year, value.month, 1)
    return None


def clean_ticker(value):
    if not value:
        return None
    return str(value).replace("$", "").strip().upper()


def as_num(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def add_title(ws, title, subtitle=None, last_col=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
    cell.font = Font(color=COLORS["white"], bold=True, size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        sub = ws.cell(2, 1, subtitle)
        sub.fill = PatternFill("solid", fgColor=COLORS["pale_blue"])
        sub.font = Font(color=COLORS["dark_gray"], italic=True)
        sub.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 34


def style_header(row):
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=medium_blue, bottom=medium_blue)


def style_table(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
            cell.alignment = Alignment(vertical="center")
    style_header(ws[min_row][min_col - 1 : max_col])


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_table(ws, name, ref):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_dropdown(ws, range_ref, values):
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(range_ref)


def load_prior_trades():
    wb_formula = load_workbook(SOURCE, data_only=False)
    wb_values = load_workbook(SOURCE, data_only=True)
    trades = []

    ws_f = wb_formula["Call  Puts"]
    ws_v = wb_values["Call  Puts"]
    for row in range(3, ws_v.max_row + 1):
        trade_no = ws_v.cell(row, 2).value
        open_date = ws_v.cell(row, 3).value
        ticker = clean_ticker(ws_v.cell(row, 4).value)
        qty = as_num(ws_v.cell(row, 5).value)
        action = ws_v.cell(row, 6).value
        opt_type = ws_v.cell(row, 7).value
        strike = as_num(ws_v.cell(row, 8).value)
        exp = ws_v.cell(row, 9).value
        entry = as_num(ws_v.cell(row, 10).value)
        close_date = ws_v.cell(row, 11).value
        rolled = ws_v.cell(row, 14).value
        close_price = as_num(ws_v.cell(row, 15).value)
        realized = as_num(ws_v.cell(row, 16).value)
        if not ticker or not qty or not action or not opt_type:
            continue

        if close_date and not close_price and action == "Sell":
            close_price = 0
        if action == "Buy" and entry is None and close_price is not None and realized is not None:
            entry = close_price - (realized / (qty * 100))

        if action == "Sell" and opt_type == "Call":
            strategy = "Covered Call"
        elif action == "Sell" and opt_type == "Put":
            strategy = "Cash Secured Put"
        elif action == "Buy":
            strategy = "Long Call/Put"
        else:
            strategy = "Other Option"

        trades.append(
            {
                "trade_id": int(trade_no) if isinstance(trade_no, (int, float)) else trade_no,
                "strategy": strategy,
                "status": "Closed" if close_date else "Open",
                "open_date": open_date,
                "close_date": close_date,
                "ticker": ticker,
                "asset": "Option",
                "action": action,
                "option_type": opt_type,
                "strike": strike,
                "expiration": exp,
                "qty": qty,
                "entry_price": entry,
                "exit_price": close_price,
                "fees": 0,
                "manual_pl": realized,
                "notes": f"Migrated from old Call/Puts tab. Rolled/Early close: {rolled or 'No/blank'}",
                "source": "Old tracker",
            }
        )

    ws_spread = wb_values["Call Debit Spreads"]
    next_id = max([t["trade_id"] for t in trades if isinstance(t["trade_id"], int)] or [0]) + 1
    for row in range(2, ws_spread.max_row + 1):
        ticker = clean_ticker(ws_spread.cell(row, 4).value)
        qty = as_num(ws_spread.cell(row, 5).value)
        cost = as_num(ws_spread.cell(row, 6).value)
        if not ticker or not qty:
            continue
        close_date = ws_spread.cell(row, 11).value
        profit = as_num(ws_spread.cell(row, 12).value)
        trades.append(
            {
                "trade_id": next_id,
                "strategy": "Debit Spread",
                "status": "Closed" if close_date else "Open",
                "open_date": ws_spread.cell(row, 3).value,
                "close_date": close_date,
                "ticker": ticker,
                "asset": "Option Spread",
                "action": "Buy",
                "option_type": "Call",
                "strike": None,
                "expiration": ws_spread.cell(row, 7).value,
                "qty": qty,
                "entry_price": cost,
                "exit_price": None,
                "fees": 0,
                "manual_pl": profit,
                "notes": "Migrated from old Call Debit Spreads tab.",
                "source": "Old tracker",
            }
        )
        next_id += 1

    return trades


def build_workbook():
    trades = CURRENT_OPTION_POSITIONS + CURRENT_STOCK_TRADES
    wb = Workbook()
    wb.remove(wb.active)
    ws_dash = wb.create_sheet("Dashboard")
    ws_guide = wb.create_sheet("How to Use")
    ws_log = wb.create_sheet("Trade Log")
    ws_options = wb.create_sheet("Options Positions")
    ws_rev = wb.create_sheet("Revenue Tracker")
    ws_pos = wb.create_sheet("Portfolio")
    ws_ideas = wb.create_sheet("Trade Ideas")
    ws_checks = wb.create_sheet("Checks")
    ws_lists = wb.create_sheet("Lists")
    ws_lists.sheet_state = "hidden"

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    build_guide(ws_guide)
    build_trade_log(ws_log, trades)
    build_options_positions(ws_options, CURRENT_OPTION_POSITIONS)
    build_revenue_tracker(ws_rev)
    build_portfolio(ws_pos, CURRENT_STOCK_POSITIONS)
    build_trade_ideas(ws_ideas)
    build_dashboard(ws_dash)
    build_checks(ws_checks)
    build_lists(ws_lists)

    for ws in [ws_dash, ws_guide, ws_log, ws_options, ws_rev, ws_pos, ws_ideas, ws_checks]:
        ws.freeze_panes = "A4"

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_FILE)


def build_trade_log(ws, trades):
    add_title(
        ws,
        "Options & Stock Trade Log",
        "One row per trade. Blue columns are the usual inputs; calculated P/L, category, collateral, and annualized return update from the row.",
        23,
    )
    headers = [
        "Trade ID",
        "Strategy",
        "Status",
        "Open Date",
        "Close Date",
        "Realization Month",
        "Ticker",
        "Asset",
        "Action",
        "Option Type",
        "Strike",
        "Expiration",
        "Qty",
        "Entry Price",
        "Exit/Close Price",
        "Fees",
        "Trade P/L / Cash Flow",
        "Realized P/L",
        "Revenue Category",
        "Collateral / Notional",
        "Annualized Return",
        "Notes",
        "Source",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws[4])

    for idx in range(MAX_TRADE_ROWS):
        row = 5 + idx
        if idx < len(trades):
            t = trades[idx]
            values = [
                t["trade_id"],
                t["strategy"],
                t["status"],
                t["open_date"],
                t["close_date"],
                None,
                t["ticker"],
                t["asset"],
                t["action"],
                t["option_type"],
                t["strike"],
                t["expiration"],
                t["qty"],
                t["entry_price"],
                t["exit_price"],
                t["fees"],
                None,
                None,
                None,
                None,
                None,
                t["notes"],
                t["source"],
            ]
            ws.append(values)
        else:
            ws.append([None] * len(headers))

        ws.cell(row, 6).value = f'=IF($E{row}="","",DATE(YEAR($E{row}),MONTH($E{row}),1))'
        ws.cell(row, 17).value = (
            f'=IF($A{row}="","",IF($H{row}="Stock",'
            f'IF($I{row}="Buy",-$M{row}*$N{row}-$P{row},IF($I{row}="Sell",IF($N{row}="", $M{row}*$O{row}-$P{row},($O{row}-$N{row})*$M{row}-$P{row}),"")),'
            f'IF($I{row}="Sell",($N{row}-IF($O{row}="",0,$O{row}))*$M{row}*100-$P{row},'
            f'IF($I{row}="Buy",(IF($O{row}="",0,$O{row})-$N{row})*$M{row}*100-$P{row},""))))'
        )
        ws.cell(row, 18).value = f'=IF($A{row}="","",IF($C{row}="Open","",$Q{row}))'
        ws.cell(row, 19).value = (
            f'=IF($A{row}="","",IFS($B{row}="Covered Call","Covered Call Premium",'
            f'$B{row}="Cash Secured Put","Cash Secured Put Premium",'
            f'$B{row}="Long Call/Put","Long Call/Put P/L",'
            f'$B{row}="Debit Spread","Long Call/Put P/L",'
            f'$B{row}="Stock Sale","Stock Sales",'
            f'TRUE,"Other"))'
        )
        ws.cell(row, 20).value = (
            f'=IF($A{row}="","",IF($B{row}="Cash Secured Put",$K{row}*$M{row}*100,'
            f'IF($B{row}="Covered Call",$K{row}*$M{row}*100,IF($H{row}="Stock",$M{row}*$N{row},""))))'
        )
        ws.cell(row, 21).value = (
            f'=IF(OR($A{row}="",$C{row}<>"Closed",$D{row}="",$E{row}="",$T{row}=0),"",'
            f'$R{row}/$T{row}*365/MAX(1,$E{row}-$D{row}))'
        )

    add_table(ws, "TradeLog", f"A4:W{4 + MAX_TRADE_ROWS}")
    style_table(ws, 4, 4 + MAX_TRADE_ROWS, 1, len(headers))
    set_widths(
        ws,
        {
            "A": 10,
            "B": 20,
            "C": 11,
            "D": 12,
            "E": 12,
            "F": 14,
            "G": 10,
            "H": 13,
            "I": 10,
            "J": 11,
            "K": 10,
            "L": 12,
            "M": 8,
            "N": 11,
            "O": 13,
            "P": 9,
            "Q": 14,
            "R": 12,
            "S": 22,
            "T": 17,
            "U": 14,
            "V": 42,
            "W": 14,
        },
    )
    for col in ["D", "E", "F", "L"]:
        for cell in ws[f"{col}5:{col}{4 + MAX_TRADE_ROWS}"]:
            cell[0].number_format = "yyyy-mm-dd"
    for col in ["K", "N", "O"]:
        for cell in ws[f"{col}5:{col}{4 + MAX_TRADE_ROWS}"]:
            cell[0].number_format = "$0.00"
    for col in ["P", "Q", "R", "T"]:
        for cell in ws[f"{col}5:{col}{4 + MAX_TRADE_ROWS}"]:
            cell[0].number_format = "$#,##0;[Red]($#,##0);-"
    for cell in ws[f"U5:U{4 + MAX_TRADE_ROWS}"]:
        cell[0].number_format = "0.0%;[Red](0.0%);-"
    for col in [2, 3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 22]:
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=5, max_row=4 + MAX_TRADE_ROWS):
            for c in cell:
                c.font = Font(color="0000FF")
    add_dropdown(ws, f"B5:B{4 + MAX_TRADE_ROWS}", ["Covered Call", "Cash Secured Put", "Long Call/Put", "Debit Spread", "Stock Sale", "Stock Buy", "Other"])
    add_dropdown(ws, f"C5:C{4 + MAX_TRADE_ROWS}", ["Open", "Closed", "Expired", "Assigned", "Rolled", "Watch"])
    add_dropdown(ws, f"H5:H{4 + MAX_TRADE_ROWS}", ["Stock", "Option", "Option Spread", "Cash"])
    add_dropdown(ws, f"I5:I{4 + MAX_TRADE_ROWS}", ["Buy", "Sell", "Expire", "Assign"])
    add_dropdown(ws, f"J5:J{4 + MAX_TRADE_ROWS}", ["Call", "Put", ""])
    ws.conditional_formatting.add(f"R5:R{4 + MAX_TRADE_ROWS}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["pale_red"])))
    ws.conditional_formatting.add(f"R5:R{4 + MAX_TRADE_ROWS}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["pale_green"])))


def build_options_positions(ws, options):
    options = [opt for opt in options if opt.get("status") == "Open"]
    add_title(
        ws,
        "Current Options Positions",
        "Open option positions from Robinhood screenshots. Positive quantity/market value indicates long exposure; negative market value indicates short option liability.",
        18,
    )
    headers = [
        "Ticker",
        "Structure",
        "Long/Short",
        "Contracts / Qty",
        "Option Type",
        "Strike / Lower Strike",
        "Upper Strike",
        "Expiration",
        "Open Date",
        "Avg Credit / Cost",
        "Current Price",
        "Original Premium / Cost",
        "Current Market Value",
        "Unrealized P/L",
        "Cash-Secured Obligation",
        "Covered Shares Used",
        "Notes",
        "Source",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws[4])
    spread_upper = {
        ("GRAB", datetime(2028, 1, 21), 10.00): 12.00,
        ("GRAB", datetime(2027, 1, 15), 7.50): 10.00,
    }
    for idx, opt in enumerate(options, start=5):
        is_short = opt["action"] == "Sell"
        structure = "Call Debit Spread" if opt["strategy"] == "Debit Spread" else f'{opt["option_type"]} {opt["strike"]:g}'
        signed_qty = -opt["qty"] if is_short else opt["qty"]
        upper = spread_upper.get((opt["ticker"], opt["expiration"], opt["strike"]))
        ws.append(
            [
                opt["ticker"],
                structure,
                "Short" if is_short else "Long",
                signed_qty,
                opt["option_type"],
                opt["strike"],
                upper,
                opt["expiration"],
                opt["open_date"],
                opt["entry_price"],
                opt["exit_price"],
                None,
                None,
                None,
                None,
                None,
                opt["notes"],
                opt["source"],
            ]
        )
        row = idx
        ws.cell(row, 12).value = f'=IF($A{row}="","",ABS($D{row})*$J{row}*100)'
        ws.cell(row, 13).value = f'=IF($A{row}="","",IF($C{row}="Short",-ABS($D{row})*$K{row}*100,ABS($D{row})*$K{row}*100))'
        ws.cell(row, 14).value = f'=IF($A{row}="","",IF($C{row}="Short",($J{row}-$K{row})*ABS($D{row})*100,($K{row}-$J{row})*ABS($D{row})*100))'
        ws.cell(row, 15).value = f'=IF(AND($C{row}="Short",$E{row}="Put"),ABS($D{row})*$F{row}*100,0)'
        if opt["strategy"] == "Covered Call":
            ws.cell(row, 16).value = f'=IF($A{row}="",0,ABS($D{row})*100)'
        else:
            ws.cell(row, 16).value = f'=IF($A{row}="",0,0)'
    max_row = max(5, 4 + len(options))
    add_table(ws, "CurrentOptions", f"A4:R{max_row}")
    style_table(ws, 4, max_row, 1, 18)
    set_widths(ws, {"A": 10, "B": 20, "C": 11, "D": 14, "E": 11, "F": 16, "G": 12, "H": 12, "I": 12, "J": 16, "K": 13, "L": 18, "M": 19, "N": 15, "O": 22, "P": 18, "Q": 48, "R": 22})
    for row in range(5, max_row + 1):
        ws.cell(row, 8).number_format = "yyyy-mm-dd"
        ws.cell(row, 9).number_format = "yyyy-mm-dd"
        for col in [6, 7, 10, 11]:
            ws.cell(row, col).number_format = "$0.00"
        for col in [12, 13, 14, 15]:
            ws.cell(row, col).number_format = "$#,##0;[Red]($#,##0);-"
    ws.conditional_formatting.add(f"N5:N{max_row}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["pale_red"])))
    ws.conditional_formatting.add(f"N5:N{max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["pale_green"])))

    summary_row = max_row + 3
    ws.cell(summary_row, 1, "Options Exposure Summary")
    ws.cell(summary_row, 1).fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws.cell(summary_row, 1).font = Font(color=COLORS["white"], bold=True)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=6)
    summary = [
        ("Net option market value", f"=SUM(M5:M{max_row})"),
        ("Open option unrealized P/L", f"=SUM(N5:N{max_row})"),
        ("Cash-secured put obligation", f"=SUM(O5:O{max_row})"),
        ("Covered shares committed", f"=SUM(P5:P{max_row})"),
        ("Long debit spread cost basis", f'=SUMIFS(L5:L{max_row},B5:B{max_row},"Call Debit Spread")'),
        ("Long debit spread market value", f'=SUMIFS(M5:M{max_row},B5:B{max_row},"Call Debit Spread")'),
    ]
    for i, (label, formula) in enumerate(summary, start=summary_row + 1):
        ws.cell(i, 1, label)
        ws.cell(i, 2, formula)
        ws.cell(i, 2).number_format = "$#,##0;[Red]($#,##0);-"
        if "shares" in label:
            ws.cell(i, 2).number_format = "#,##0"


def build_guide(ws):
    add_title(
        ws,
        "How to Use This Tracker",
        "Designed for a Google Sheets or Excel workflow: one trade row, formula summaries, manual current positions, and a running idea log.",
        8,
    )
    sections = [
        (
            "1. Log each trade",
            "Use Trade Log for options and stock trades. Fill the blue input columns: strategy, status, open/close dates, ticker, asset/action, option type, strike, expiration, quantity, entry price, close price, fees, and notes. The realized P/L, category, collateral/notional, and annualized return columns calculate automatically.",
        ),
        (
            "2. Review monthly income",
            "Revenue Tracker rolls closed trades into monthly buckets: covered call premiums, cash secured put premiums, long option/spread P/L, stock sales, and other. The ticker summary underneath shows which symbols are producing or losing money.",
        ),
        (
            "3. Maintain portfolio weights",
            "Portfolio is intentionally manual for current shares, prices, and average cost. Enter current broker balances there; market value, weight, unrealized P/L, open covered calls, open short puts, and cash-secured obligation update from formulas.",
        ),
        (
            "4. Train future recommendations",
            "Use Trade Ideas to capture sample setups, combinations, thesis, max profit/loss, result, and lessons. Over time, this becomes the preference file for what kinds of option structures you like, avoid, or want me to compare.",
        ),
        (
            "5. Updating with Codex",
            "You can paste trade confirmations, broker export rows, or plain-English trade notes and ask me to append them. Best input format is: date, ticker, strategy, buy/sell, call/put, strike, expiration, quantity, entry/exit price, fees, status, and any assignment/roll notes.",
        ),
    ]
    row = 4
    for title, body in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1, title)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=COLORS["navy"])
        ws.cell(row, 1).font = Font(color=COLORS["white"], bold=True)
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=8)
        ws.cell(row, 1, body)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=COLORS["gray"])
        ws.cell(row, 1).border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
        ws.row_dimensions[row].height = 36
        ws.row_dimensions[row + 1].height = 18
        row += 3
    set_widths(ws, {"A": 18, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16})


def build_revenue_tracker(ws):
    add_title(ws, "Monthly Revenue Tracker", "Formula-driven view of realized gains/losses by month and income bucket.", 9)
    headers = [
        "Month",
        "Covered Call Premium",
        "Cash Secured Put Premium",
        "Long Call/Put P/L",
        "Stock Sales",
        "Other",
        "Total Realized P/L",
        "Trade Count",
        "Avg P/L / Trade",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws[4])
    start = datetime(TRACKING_YEAR, 1, 1)
    for i in range(24):
        row = 5 + i
        month = datetime(start.year + (start.month - 1 + i) // 12, (start.month - 1 + i) % 12 + 1, 1)
        ws.cell(row, 1, month)
        categories = [
            "Covered Call Premium",
            "Cash Secured Put Premium",
            "Long Call/Put P/L",
            "Stock Sales",
            "Other",
        ]
        for idx, cat in enumerate(categories, start=2):
            ws.cell(row, idx).value = f'=SUMIFS(\'Trade Log\'!$R$5:$R$1004,\'Trade Log\'!$F$5:$F$1004,$A{row},\'Trade Log\'!$S$5:$S$1004,{get_column_letter(idx)}$4)'
        ws.cell(row, 7).value = f"=SUM(B{row}:F{row})"
        ws.cell(row, 8).value = f'=COUNTIFS(\'Trade Log\'!$F$5:$F$1004,$A{row},\'Trade Log\'!$R$5:$R$1004,"<>")'
        ws.cell(row, 9).value = f'=IF($H{row}=0,"",$G{row}/$H{row})'
    add_table(ws, "MonthlyRevenue", "A4:I28")
    style_table(ws, 4, 28, 1, 9)
    set_widths(ws, {"A": 13, "B": 20, "C": 22, "D": 18, "E": 14, "F": 12, "G": 18, "H": 12, "I": 15})
    for row in range(5, 29):
        ws.cell(row, 1).number_format = "mmm yyyy"
        for col in range(2, 8):
            ws.cell(row, col).number_format = "$#,##0;[Red]($#,##0);-"
        ws.cell(row, 9).number_format = "$#,##0;[Red]($#,##0);-"

    ticker_header_row = 32
    ws.cell(ticker_header_row, 1, "Ticker")
    for i, label in enumerate(["Covered Call", "Cash Secured Put", "Long Option / Spread", "Stock Sales", "Other", "Total"], start=2):
        ws.cell(ticker_header_row, i, label)
    style_header(ws[ticker_header_row][0:7])
    for i in range(50):
        row = ticker_header_row + 1 + i
        ws.cell(row, 1).value = f'=IFERROR(INDEX(SORT(UNIQUE(FILTER(\'Trade Log\'!$G$5:$G$1004,\'Trade Log\'!$G$5:$G$1004<>""))),ROW()-{ticker_header_row}),"")'
        cats = ["Covered Call Premium", "Cash Secured Put Premium", "Long Call/Put P/L", "Stock Sales", "Other"]
        for idx, cat in enumerate(cats, start=2):
            ws.cell(row, idx).value = f'=IF($A{row}="","",SUMIFS(\'Trade Log\'!$R$5:$R$1004,\'Trade Log\'!$G$5:$G$1004,$A{row},\'Trade Log\'!$S$5:$S$1004,"{cat}"))'
        ws.cell(row, 7).value = f'=IF($A{row}="","",SUM(B{row}:F{row}))'
        for col in range(2, 8):
            ws.cell(row, col).number_format = "$#,##0;[Red]($#,##0);-"
    style_table(ws, ticker_header_row, ticker_header_row + 50, 1, 7)

    chart = BarChart()
    chart.title = "Monthly Realized P/L by Bucket"
    chart.y_axis.title = "Realized P/L"
    chart.x_axis.title = "Month"
    data = Reference(ws, min_col=2, max_col=6, min_row=4, max_row=28)
    cats = Reference(ws, min_col=1, min_row=5, max_row=28)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "K4")


def build_portfolio(ws, positions):
    add_title(ws, "Portfolio Positions & Weights", "Current stock positions from screenshots. Weights, unrealized P/L, and option exposure update automatically.", 14)
    headers = [
        "Ticker",
        "Shares",
        "Current Price",
        "Avg Cost",
        "Market Value",
        "Cost Basis",
        "Unrealized P/L",
        "Weight",
        "Open CC Contracts",
        "Open CSP Contracts",
        "Cash Secured Obligation",
        "Covered Shares",
        "Coverage %",
        "Notes",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws[4])
    for i in range(MAX_POSITION_ROWS):
        row = 5 + i
        ws.append([None] * len(headers))
        if i < len(positions):
            pos = positions[i]
            ws.cell(row, 1, pos["ticker"])
            ws.cell(row, 2, pos["shares"])
            ws.cell(row, 3, pos["price"])
            ws.cell(row, 4, pos["avg_cost"])
            ws.cell(row, 14, pos["notes"])
        ws.cell(row, 5).value = f'=IF($A{row}="","",IFERROR($B{row}*$C{row},0))'
        ws.cell(row, 6).value = f'=IF($A{row}="","",IFERROR($B{row}*$D{row},0))'
        ws.cell(row, 7).value = f'=IF($A{row}="","",$E{row}-$F{row})'
        ws.cell(row, 8).value = f'=IF($A{row}="","",IFERROR($E{row}/SUM($E$5:$E${4 + MAX_POSITION_ROWS}),0))'
        ws.cell(row, 9).value = f'=IF($A{row}="","",SUMIFS(\'Trade Log\'!$M$5:$M$1004,\'Trade Log\'!$G$5:$G$1004,$A{row},\'Trade Log\'!$B$5:$B$1004,"Covered Call",\'Trade Log\'!$C$5:$C$1004,"Open"))'
        ws.cell(row, 10).value = f'=IF($A{row}="","",SUMIFS(\'Trade Log\'!$M$5:$M$1004,\'Trade Log\'!$G$5:$G$1004,$A{row},\'Trade Log\'!$B$5:$B$1004,"Cash Secured Put",\'Trade Log\'!$C$5:$C$1004,"Open"))'
        ws.cell(row, 11).value = f'=IF($A{row}="","",SUMIFS(\'Trade Log\'!$T$5:$T$1004,\'Trade Log\'!$G$5:$G$1004,$A{row},\'Trade Log\'!$B$5:$B$1004,"Cash Secured Put",\'Trade Log\'!$C$5:$C$1004,"Open"))'
        ws.cell(row, 12).value = f'=IF($A{row}="","",$I{row}*100)'
        ws.cell(row, 13).value = f'=IFERROR($L{row}/$B{row},0)'
    add_table(ws, "PortfolioPositions", f"A4:N{4 + MAX_POSITION_ROWS}")
    style_table(ws, 4, 4 + MAX_POSITION_ROWS, 1, len(headers))
    set_widths(ws, {"A": 10, "B": 10, "C": 13, "D": 12, "E": 15, "F": 14, "G": 14, "H": 10, "I": 15, "J": 16, "K": 21, "L": 14, "M": 12, "N": 28})
    for col in ["C", "D"]:
        for row in range(5, 5 + MAX_POSITION_ROWS):
            ws[f"{col}{row}"].number_format = "$0.00"
    for col in ["E", "F", "G", "K"]:
        for row in range(5, 5 + MAX_POSITION_ROWS):
            ws[f"{col}{row}"].number_format = "$#,##0;[Red]($#,##0);-"
    for col in ["H", "M"]:
        for row in range(5, 5 + MAX_POSITION_ROWS):
            ws[f"{col}{row}"].number_format = "0.0%;[Red](0.0%);-"
    for col in ["B", "C", "D", "N"]:
        for row in range(5, 5 + MAX_POSITION_ROWS):
            ws[f"{col}{row}"].font = Font(color="0000FF")
    ws.conditional_formatting.add(f"G5:G{4 + MAX_POSITION_ROWS}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["pale_red"])))
    ws.conditional_formatting.add(f"G5:G{4 + MAX_POSITION_ROWS}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=COLORS["pale_green"])))

    pie = PieChart()
    pie.title = "Portfolio Weight"
    data = Reference(ws, min_col=5, min_row=4, max_row=min(20, 4 + len(positions)))
    labels = Reference(ws, min_col=1, min_row=5, max_row=min(20, 4 + len(positions)))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 8
    pie.width = 12
    ws.add_chart(pie, "P4")


def build_trade_ideas(ws):
    add_title(
        ws,
        "Trade Ideas & Recommendation Log",
        "Use this to feed examples over time: proposed setup, logic, risks, outcome, and what you want me to remember for future recommendations.",
        15,
    )
    headers = [
        "Idea ID",
        "Date",
        "Ticker",
        "Market View",
        "Setup / Strategy",
        "Legs",
        "Net Credit / Debit",
        "Max Profit",
        "Max Loss",
        "Breakeven",
        "Target",
        "Status",
        "Outcome P/L",
        "Lesson / Preference",
        "Link to Trade ID",
    ]
    ws.append([])
    ws.append(headers)
    style_header(ws[4])
    for i in range(200):
        ws.append([None] * len(headers))
    add_table(ws, "TradeIdeas", "A4:O204")
    style_table(ws, 4, 204, 1, len(headers))
    set_widths(ws, {"A": 10, "B": 12, "C": 10, "D": 20, "E": 24, "F": 34, "G": 16, "H": 13, "I": 13, "J": 13, "K": 13, "L": 12, "M": 13, "N": 34, "O": 14})
    for col in ["B"]:
        for row in range(5, 205):
            ws[f"{col}{row}"].number_format = "yyyy-mm-dd"
    for col in ["G", "H", "I", "J", "K", "M"]:
        for row in range(5, 205):
            ws[f"{col}{row}"].number_format = "$#,##0;[Red]($#,##0);-"
    add_dropdown(ws, "L5:L204", ["Researching", "Proposed", "Entered", "Passed", "Closed", "Archived"])
    add_dropdown(ws, "D5:D204", ["Bullish", "Neutral-Bullish", "Neutral", "Neutral-Bearish", "Bearish", "Volatility"])


def build_dashboard(ws):
    add_title(ws, "Options Income Dashboard", "High-level view across realized P/L, strategy buckets, current portfolio weight, and open option exposure.", 10)
    labels = [
        ("A4", "YTD Realized P/L", "B4", '=SUM(\'Trade Log\'!$R$5:$R$1004)'),
        ("D4", "Avg Monthly P/L", "E4", '=IFERROR(AVERAGEIF(\'Revenue Tracker\'!$G$5:$G$28,"<>0",\'Revenue Tracker\'!$G$5:$G$28),0)'),
        ("G4", "Portfolio Value", "H4", '=SUM(Portfolio!$E$5:$E$154)'),
        ("A7", "Open CSP Collateral", "B7", '=SUMIFS(\'Trade Log\'!$T$5:$T$1004,\'Trade Log\'!$B$5:$B$1004,"Cash Secured Put",\'Trade Log\'!$C$5:$C$1004,"Open")'),
        ("D7", "Open CC Contracts", "E7", '=SUMIFS(\'Trade Log\'!$M$5:$M$1004,\'Trade Log\'!$B$5:$B$1004,"Covered Call",\'Trade Log\'!$C$5:$C$1004,"Open")'),
        ("G7", "Closed Trade Count", "H7", '=COUNTIF(\'Trade Log\'!$C$5:$C$1004,"Closed")'),
    ]
    for label_cell, label, value_cell, formula in labels:
        ws[label_cell] = label
        ws[label_cell].fill = PatternFill("solid", fgColor=COLORS["pale_blue"])
        ws[label_cell].font = Font(bold=True, color=COLORS["dark_gray"])
        ws[value_cell] = formula
        ws[value_cell].fill = PatternFill("solid", fgColor=COLORS["white"])
        ws[value_cell].font = Font(bold=True, size=13)
        ws[value_cell].border = Border(bottom=medium_blue)
    for cell in ["B4", "E4", "H4", "B7"]:
        ws[cell].number_format = "$#,##0;[Red]($#,##0);-"
    ws["E7"].number_format = "#,##0"
    ws["H7"].number_format = "#,##0"

    ws["A11"] = "Recent Monthly Performance"
    ws["A11"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A11"].font = Font(color=COLORS["white"], bold=True)
    ws.merge_cells("A11:I11")
    for row in range(12, 18):
        for col in range(1, 10):
            target = ws.cell(row, col)
            target.value = f"='Revenue Tracker'!{get_column_letter(col)}{row-7}"
            if col == 1:
                target.number_format = "mmm yyyy"
            elif col in range(2, 8) or col == 9:
                target.number_format = "$#,##0;[Red]($#,##0);-"
    style_table(ws, 12, 17, 1, 9)
    set_widths(ws, {"A": 16, "B": 15, "C": 17, "D": 15, "E": 14, "F": 12, "G": 15, "H": 12, "I": 14})


def build_checks(ws):
    add_title(ws, "Workbook Checks", "Quick checks for missing inputs and possible formula issues.", 7)
    headers = ["Check", "Actual", "Expected", "Difference", "Status", "Notes", "Fix"]
    ws.append([])
    ws.append(headers)
    style_header(ws[4])
    rows = [
        ("Trade rows missing ticker", '=COUNTIFS(\'Trade Log\'!$A$5:$A$1004,"<>",\'Trade Log\'!$G$5:$G$1004,"")', 0, "=B5-C5", '=IF(D5=0,"OK","Review")', "Every trade should have a ticker.", "Fill Trade Log ticker."),
        ("Closed trades missing close date", '=COUNTIFS(\'Trade Log\'!$C$5:$C$1004,"Closed",\'Trade Log\'!$E$5:$E$1004,"")', 0, "=B6-C6", '=IF(D6=0,"OK","Review")', "Closed trades need a close date for monthly reporting.", "Add close date."),
        ("Open CSP collateral missing", '=COUNTIFS(\'Trade Log\'!$B$5:$B$1004,"Cash Secured Put",\'Trade Log\'!$C$5:$C$1004,"Open",\'Trade Log\'!$T$5:$T$1004,0)', 0, "=B7-C7", '=IF(D7=0,"OK","Review")', "Open CSP rows should calculate collateral.", "Check strike and qty."),
        ("Portfolio rows missing price", '=COUNTIFS(Portfolio!$A$5:$A$154,"<>",Portfolio!$C$5:$C$154,"")', 0, "=B8-C8", '=IF(D8=0,"OK","Review")', "Current prices drive weights.", "Enter price or remove unused ticker."),
        ("Overall status", '=COUNTIF($E$5:$E$8,"Review")', 0, "=B9-C9", '=IF(B9=0,"OK","Review")', "Rolls up checks above.", "Resolve Review rows."),
    ]
    for row in rows:
        ws.append(list(row))
    style_table(ws, 4, 9, 1, 7)
    set_widths(ws, {"A": 28, "B": 12, "C": 12, "D": 12, "E": 12, "F": 42, "G": 28})
    ws.conditional_formatting.add("E5:E9", FormulaRule(formula=['E5="OK"'], fill=PatternFill("solid", fgColor=COLORS["pale_green"])))
    ws.conditional_formatting.add("E5:E9", FormulaRule(formula=['E5="Review"'], fill=PatternFill("solid", fgColor=COLORS["pale_orange"])))


def build_lists(ws):
    lists = {
        "Strategies": ["Covered Call", "Cash Secured Put", "Long Call/Put", "Debit Spread", "Stock Sale", "Stock Buy", "Other"],
        "Statuses": ["Open", "Closed", "Expired", "Assigned", "Rolled", "Watch"],
        "Assets": ["Stock", "Option", "Option Spread", "Cash"],
        "Actions": ["Buy", "Sell", "Expire", "Assign"],
        "Option Types": ["Call", "Put"],
        "Idea Statuses": ["Researching", "Proposed", "Entered", "Passed", "Closed", "Archived"],
    }
    for col_idx, (name, values) in enumerate(lists.items(), start=1):
        ws.cell(1, col_idx, name)
        ws.cell(1, col_idx).font = Font(bold=True)
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row_idx, col_idx, value)


if __name__ == "__main__":
    build_workbook()
    print(OUT_FILE)
